#!/usr/bin/env python3
"""Build SmartByteKC competitor price-scouting workbook (SMART-HOME / LOW-VOLTAGE)."""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---- Palette (SmartByte brand: dark slate + electric emerald) ----
NAVY   = "0B0F19"
GREEN  = "10B981"
LGREEN = "D7F0DD"
BLUE   = "2E6FE0"
LBLUE  = "DCE7FB"
GREY   = "F2F4F7"
LAMBER = "FCEFD2"
WHITE  = "FFFFFF"
DARK   = "222222"

thin = Side(style="thin", color="C9CFDA")
border = Border(left=thin, right=thin, top=thin, bottom=thin)
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center", wrap_text=True)

def hdr(cell, text, fill=GREEN, color=WHITE, size=11):
    cell.value = text
    cell.font = Font(bold=True, color=color, size=size)
    cell.fill = PatternFill("solid", fgColor=fill)
    cell.alignment = center
    cell.border = border

def title_row(ws, text, span, row=1):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=span)
    c = ws.cell(row=row, column=1, value=text)
    c.font = Font(bold=True, color=WHITE, size=14)
    c.fill = PatternFill("solid", fgColor=NAVY)
    c.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 26

def setw(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

wb = Workbook()

# =====================================================================
# SHEET 1 — Competitor Directory (smart-home / low-voltage, KC metro)
# =====================================================================
ws = wb.active
ws.title = "Competitor Directory"
title_row(ws, "SmartByteKC — KC Smart-Home / Low-Voltage Competitor Directory  (Week 1 Foundation)", 7)
ws.cell(row=2, column=1, value="Status legend:  PUBLISHED = confirmed pricing on their site  |  REVIEW = customer review  |  BENCH = regional market estimate").font = Font(italic=True, size=9, color="555555")
ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=7)

headers = ["Competitor", "Website", "Phone", "Location", "Years / Type", "Known Pricing Signals", "Notes"]
for i, h in enumerate(headers, start=1):
    hdr(ws.cell(row=3, column=i), h)

competitors = [
    ["Bishop Low Voltage Solutions LLC", "bishoplowvoltage.com", "(913) 674-6625", "KC metro, KS side",
     "Local low-voltage specialist; home automation + security",
     "Free estimate; no published rate card (BENCH).",
     "Closest pure-play low-voltage competitor. Local, relationship-driven."],
    ["Simplicity", "simplicitykc.com", "(866) 691-7754", "19535 Metcalf Ave, Stilwell KS 66085",
     "Smart-home automation design + install",
     "Project-based quotes; no rate card (BENCH).",
     "Design-forward integrator. Mid-to-high-end positioning."],
    ["Kinkade Home Theater", "kinkadeht.com", "—", "Lenexa, KS",
     "Elite home theater + low-voltage",
     "Custom quotes; premium tier (BENCH).",
     "High-end media rooms. Not a price anchor."],
    ["Wired By Design", "wiredby.design", "—", "KC metro",
     "Networking / structured cabling",
     "Cat6 ~$150-300/drop; fiber $300-800+/run (PUBLISHED).",
     "Transparent per-drop pricing — strong benchmark for wiring."],
    ["KC Dynamic Wiring", "kcdynamicwiring.com", "—", "KC metro",
     "WiFi optimization, mesh, structured wiring",
     "Free estimate; competitive Cat6/fiber rates (BENCH).",
     "Direct wiring competitor; speed/coverage angle."],
    ["Homedia Solutions", "—", "—", "KC metro",
     "Full-service smart-home integrator",
     "Min project fee $2,500+; 4-8 wk lead (REVIEW).",
     "Full integrator. High barrier = our opening for smaller jobs."],
    ["MVP Electric, Heating & Cooling", "themvpkc.com", "(913) 210-5512", "6820 Squibb Rd, Mission KS",
     "Multi-trade (HVAC/plumbing/low-voltage)",
     "Upfront pricing policy; no published rate card (BENCH).",
     "Bundled home-services play, not a specialist."],
    ["Vivint", "vivint.com", "—", "National (KC presence)",
     "Security-first smart-home",
     "Install ~$225 w/ 2-yr warranty; monitoring sub (PUBLISHED).",
     "Security-first; contract/subscription model."],
    ["Shield Security Systems", "—", "—", "KC metro",
     "Security-first provider",
     "Quote-based; monitoring plans (BENCH).",
     "Security angle; less whole-home automation."],
    ["Best Buy Geek Squad", "bestbuy.com", "—", "Multiple KC stores",
     "Retail tech install",
     "Smart thermostat $199 flat; smart lock $119-185 (PUBLISHED).",
     "DIY-helper price anchor; no custom design."],
]
r = 4
for row in competitors:
    for i, val in enumerate(row, start=1):
        c = ws.cell(row=r, column=i, value=val)
        c.border = border
        c.alignment = wrap
        if i == 1:
            c.font = Font(bold=True, color=NAVY)
        if r % 2 == 0:
            c.fill = PatternFill("solid", fgColor=GREY)
    r += 1
setw(ws, [30, 24, 18, 28, 34, 40, 38])
ws.freeze_panes = "A4"

# =====================================================================
# SHEET 2 — Price Scouting Matrix (service x competitor)
# =====================================================================
ws2 = wb.create_sheet("Price Scouting Matrix")
comp_cols = ["Bishop", "Simplicity", "WiredByDesign", "KC Dynamic", "Homedia",
             "Vivint", "Geek Squad", "Market Benchmark"]
title_row(ws2, "Price Scouting Matrix — Service Item x Competitor  (blank = not published / to verify)", len(comp_cols)+2)
sub = ("Confidence:  PUB = published on site  |  REV = customer review  |  EST = estimate/benchmark  |  - = unknown (call to verify)\n"
       "Market Benchmark column sources: CEDIA 2026 integrator data, HomeCostCalc KS 2026, SmarthomeDeck 2026, WiredByDesign published rates, freedomtodays KC guide 2026.")
ws2.merge_cells(start_row=2, start_column=1, end_row=3, end_column=len(comp_cols)+2)
c = ws2.cell(row=2, column=1, value=sub)
c.font = Font(italic=True, size=9, color="555555"); c.alignment = wrap

hdr(ws2.cell(row=4, column=1), "Service Item")
hdr(ws2.cell(row=4, column=2), "Unit")
for j, name in enumerate(comp_cols, start=3):
    hdr(ws2.cell(row=4, column=j), name)
hdr(ws2.cell(row=4, column=len(comp_cols)+3), "Confidence")

matrix = [
    ["Site survey / consultation", "$", ["FREE", "quote", "FREE", "FREE", "quote", "quote", "—", "0-150"], "mixed"],
    ["Smart thermostat install", "fixed $", ["—", "—", "—", "—", "—", 225, 199, "119-225"], "pub"],
    ["Smart lock install", "fixed $", ["—", "—", "—", "—", "—", "—", "119-185", "119-210"], "pub"],
    ["Mesh Wi-Fi design + install", "fixed $", ["—", "—", "150-500", "150-500", "—", "—", "—", "150-500"], "mixed"],
    ["Cat6 structured drop", "$/drop", [150, 200, "150-300", "150-300", 250, "—", "—", "150-300"], "pub"],
    ["Fiber run (complex)", "$/run", ["—", "—", "300-800+", "300-800+", "—", "—", "—", "300-800"], "pub"],
    ["Security cam (2K wireless, pro)", "$/cam", ["—", "—", "—", "—", "—", 200, "—", "520-900"], "mixed"],
    ["Security cam (PoE 4K, pro)", "$/cam", ["—", "—", "—", "—", "—", "—", "—", "639-1500"], "bench"],
    ["4-camera system (pro, 2K)", "fixed $", ["—", "—", "—", "—", "—", "—", "—", "1200-2000"], "bench"],
    ["Whole-home lighting (8+ switches)", "fixed $", ["—", "—", "—", "—", 1299, "—", "—", "749-1299"], "mixed"],
    ["Full smart-home (entry tier)", "fixed $", ["—", "—", "—", "—", "2500+", "—", "—", "500-2500"], "bench"],
    ["Full smart-home (mid tier)", "fixed $", ["—", "—", "—", "—", "25000+", "—", "—", "5000-30000"], "bench"],
    ["Integrator labor rate", "$/hr", ["120", "175", "—", "—", "175", "—", "—", "120-175"], "mixed"],
]
r = 5
for row in matrix:
    svc, unit, vals, conf = row
    a = ws2.cell(row=r, column=1, value=svc); a.font = Font(bold=True, color=NAVY); a.border = border; a.alignment = wrap
    u = ws2.cell(row=r, column=2, value=unit); u.alignment = center; u.border = border
    for j, v in enumerate(vals, start=3):
        cell = ws2.cell(row=r, column=j, value=v)
        cell.alignment = center; cell.border = border
        sval = str(v)
        if sval == "—":
            cell.fill = PatternFill("solid", fgColor="FBE3E3")
        elif sval in ("FREE",):
            cell.fill = PatternFill("solid", fgColor=LGREEN)
        elif any(ch.isdigit() for ch in sval):
            cell.fill = PatternFill("solid", fgColor=LBLUE)
    cc = ws2.cell(row=r, column=len(comp_cols)+3, value=conf); cc.alignment = center; cc.border = border
    cc.font = Font(italic=True, size=9)
    r += 1
ws2.cell(row=r+1, column=1, value="Note: Full-service integrators (Homedia, Simplicity) carry $2,500+ minimums and 4-8 week leads — leaving a wide opening for SmartByte on single-room, same-week, no-minimum smart-home jobs.").font = Font(italic=True, size=9, color="555555")
ws2.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=len(comp_cols)+3)
setw(ws2, [34, 9] + [14]*len(comp_cols) + [12])
ws2.freeze_panes = "C5"

# =====================================================================
# SHEET 3 — Market Benchmarks (research summary)
# =====================================================================
ws3 = wb.create_sheet("Market Benchmarks")
title_row(ws3, "KC Smart-Home / Low-Voltage Market Benchmarks (research sources)", 5)
for i, h in enumerate(["Metric", "Low", "Avg / Typical", "High", "Source"], start=1):
    hdr(ws3.cell(row=2, column=i), h)
bench = [
    ["Integrator labor rate", "$120", "$150", "$175", "CEDIA 2026 / freedomtodays KC 2026"],
    ["Site survey / consultation", "$0", "$0", "$150", "Most local firms free; Geek Squad n/a"],
    ["Smart thermostat install", "$119", "$199", "$225", "Best Buy Geek Squad / Vivint 2026"],
    ["Smart lock install", "$119", "$185", "$210", "Best Buy / CEDIA install data 2026"],
    ["Mesh Wi-Fi design + install", "$150", "$325", "$500", "WiredByDesign / KC Dynamic 2026"],
    ["Cat6 structured drop", "$150", "$225", "$300", "WiredByDesign published rate 2026"],
    ["Fiber run (complex)", "$300", "$550", "$800+", "WiredByDesign 2026"],
    ["Security cam (2K wireless, pro)", "$520", "$710", "$900", "HomeCostCalc KS 2026"],
    ["Security cam (PoE 4K, pro)", "$639", "$1,070", "$1,500", "HomeCostCalc / tec-tel 2025"],
    ["4-camera system (pro, 2K)", "$1,200", "$1,600", "$2,000", "HomeCostCalc KS 2026"],
    ["Whole-home lighting (8+ switches)", "$749", "$1,024", "$1,299", "SmarthomeDeck 2026"],
    ["Entry smart-home project", "$500", "$1,500", "$2,500", "National Home Automation Authority 2026"],
    ["Mid-range smart-home project", "$5,000", "$17,500", "$30,000", "National Home Automation Authority 2026"],
]
r = 3
for row in bench:
    for i, v in enumerate(row, start=1):
        cell = ws3.cell(row=r, column=i, value=v); cell.border = border; cell.alignment = wrap
        if i == 1: cell.font = Font(bold=True, color=NAVY)
        if r % 2 == 0: cell.fill = PatternFill("solid", fgColor=GREY)
    r += 1
setw(ws3, [32, 12, 16, 14, 44])

# =====================================================================
# SHEET 4 — SmartByte Positioning (to fill in)
# =====================================================================
ws4 = wb.create_sheet("SmartByte Pricing Plan")
title_row(ws4, "SmartByteKC - Our Pricing Position (owner to set)", 5)
for i, h in enumerate(["Service Item", "Our Cost (labor+materials)", "Target Price", " vs Market Avg", "Margin / Notes"], start=1):
    hdr(ws4.cell(row=2, column=i), h)
plan = [
    ["Site survey / consultation", "", "FREE", "match free norm", "Lead-gen play; book jobs on-site"],
    ["Smart thermostat install", "", "e.g. $169", "below $199 avg", "Beat Geek Squad on local service"],
    ["Smart lock install", "", "e.g. $149", "below $185 avg", ""],
    ["Mesh Wi-Fi design + install", "", "e.g. $299", "below $325 avg", "High-volume entry service"],
    ["Cat6 structured drop", "", "e.g. $199", "below $225 avg", "Undercut WiredByDesign on small jobs"],
    ["Security cam (2K wireless, pro)", "", "e.g. $599", "vs $710 avg", ""],
    ["4-camera system (pro, 2K)", "", "e.g. $1,499", "below $1,600 avg", "Bundle play vs integrators"],
    ["Whole-home lighting (8+ switches)", "", "e.g. $1,099", "below $1,299", ""],
    ["Full smart-home (entry tier)", "", "e.g. $1,299", "vs $1,500 avg", "No $2,500 minimum = our wedge"],
]
r = 3
for row in plan:
    for i, v in enumerate(row, start=1):
        cell = ws4.cell(row=r, column=i, value=v); cell.border = border; cell.alignment = wrap
        if i == 1: cell.font = Font(bold=True, color=NAVY)
        if r % 2 == 0: cell.fill = PatternFill("solid", fgColor=GREY)
    r += 1
setw(ws4, [32, 24, 16, 18, 34])
ws4.cell(row=r+1, column=1, value="Strategy: undercut full-service integrators (Homedia/Simplicity $2,500+ min) and beat retail (Geek Squad) on local, same-week, no-minimum smart-home service. Aim: below market avg on every line while leading with 'we're your KC neighbor' positioning.").font = Font(italic=True, size=9, color="555555")
ws4.merge_cells(start_row=r+1, start_column=1, end_row=r+1, end_column=5)

import os
os.makedirs("C:/Users/Gibby/Documents/SmartByteKC/Finance", exist_ok=True)
out1 = "C:/Users/Gibby/Documents/SmartByteKC/Finance/PriceScout_Competitors.xlsx"
out2 = "C:/Users/Gibby/smartbytekc/sheets/PriceScout_Competitors.xlsx"
wb.save(out1)
wb.save(out2)
print("Saved:\n", out1, "\n", out2)
print("Sheets:", wb.sheetnames)
